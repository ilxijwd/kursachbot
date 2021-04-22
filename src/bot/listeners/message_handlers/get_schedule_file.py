import os
from tempfile import TemporaryFile

import openpyxl

from src.bot import tb
from src.db import session, Person, Faculty, Schedule, ScheduleMeeting, Meeting


@tb.message_handler(
    func=lambda msg: hasattr(msg.reply_to_message, 'text') and
    msg.reply_to_message.text == "Очікую файл розкладу" and
    (msg.document.mime_type == 'text/application/vnd.ms-excel' or
     msg.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
    content_types=['document']
)
def get_schedule_file(msg):
    person = session.query(Person).filter(Person.telegram_id == msg.chat.id).first()

    # FILE SIZE CHECK
    if msg.document.file_size > 500000:
        return tb.send_message(msg.chat.id, "Файл великого розміру (0.5mb max)")

    # DOWNLOADING THE FILE
    tg_cloud_file = tb.get_file(msg.document.file_id)
    binary_data = tb.download_file(tg_cloud_file.file_path)

    # CREATING TEMPORARY FILE WRAPPER
    schedule_file = TemporaryFile()
    schedule_file.write(binary_data)

    # DELETING CURRENT SCHEDULE FROM DB
    if person.head.faculty.schedule:
        session.delete(person.head.faculty.schedule)
        session.commit()

    # PARSING TABLE WITH DAYS OF PRACTICE
    schedule = Schedule(faculty=person.head.faculty)
    try:
        schedule_sheet = openpyxl.load_workbook(schedule_file).active
        for row_idx in range(2, schedule_sheet.max_row + 1):
            schedule.schedule_meetings.append(
                ScheduleMeeting(
                    schedule=schedule,
                    meeting=Meeting(
                        date=schedule_sheet.cell(row_idx, 1).value.strftime("%Y-%m-%d"),
                        start_time=schedule_sheet.cell(row_idx, 2).value.strftime("%H:%M"),
                        end_time=schedule_sheet.cell(row_idx, 3).value.strftime("%H:%M"),
                        place=schedule_sheet.cell(row_idx, 4).value,
                    )
                )
            )
    except (ValueError, TypeError):
        tb.reply_to(msg, "Помилка у парсингу файла")
        schedule_file.close()
        return

    # INSERTING NEW SCHEDULE TO DB
    person.head.faculty.schedule = schedule
    session.commit()

    # NOTIFYING FOR SUCCESS AND CLEANING UP
    tb.reply_to(msg, "Успішно завантажено новий розклад!")
    schedule_file.close()

    # INFORMING THE STUDENTS WITH NEW SCHEDULE BY SENDING THE FILE
    not_informed_students = []
    informed_students = []
    for group in person.head.faculty.groups:
        for student in group.students:
            if not student.person.telegram_id:
                not_informed_students.append(student)
                continue

            tb.send_document(
                student.person.telegram_id,
                msg.document.file_id,
                caption="Шановний студент! Розклад консультацій змінено, лови актуальний"
            )
            informed_students.append(student)

    # NOTIFYING ABOUT THE RESULTS ABOUT STUDENTS INFORMATION
    tb.send_message(
        msg.chat.id,
        f"Інформованих студентів: {len(informed_students)}\n"
        "Не отримали повідомленя: \n"
        f"{os.linesep.join((f'{i + 1}. {student.person.fullname}' for (i, student) in enumerate(not_informed_students)))}",
        parse_mode='Markdown'
    )
